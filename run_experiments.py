import itertools
import traceback
import csv
import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from helpers import get_problems, initiate_stats
from multipivot_simplex_solver import multipivot_simplex


def run_multipivot_simplex(i, kwargs):
    """
    Executes the Multi-Pivot Simplex algorithm for a specified problem instance and collects performance statistics.

    This function retrieves the problem data (A, b, C) corresponding to the specified problem index,
    initializes performance statistics, and runs the Multi-Pivot Simplex algorithm using the provided
    keyword arguments. The results include the optimal value, feasibility of the solution, maximum reduced
    cost, and updated statistics.

    Parameters:
    ----------
    i : int
        The index of the problem to be solved, which corresponds to one of the problem examples returned by `get_problems()`.
    kwargs : dict
        A dictionary of keyword arguments that are passed to the `multipivot_simplex` function. This allows for customization
        of the algorithm's behavior (e.g., number of pivots, whether cuts are used, etc.).

    Returns:
    -------
    tuple:
        - Z : float or None
            The optimal value of the objective function. Returns `None` if an exception occurs.
        - feasible : bool or None
            A boolean indicating whether the solution is feasible (all elements of X are non-negative). Returns `None` if an exception occurs.
        - max_RC : float or None
            The maximum value of the reduced costs (RC) vector. Returns `None` if an exception occurs.
        - stats : dict
            A dictionary containing performance statistics such as the number of iterations, cuts, and inversions.
            If an exception occurs, this dictionary will also include an error message.

    Example:
    --------
    # Define keyword arguments for the algorithm
    kwargs = {
        "Cut": True,
        "Order": False,
        "Pivots": 2,
        "maxiters": 100
    }

    # Run the Multi-Pivot Simplex algorithm for problem 1
    Z, feasible,  max_RC, stats = run_multipivot_simplex(1, kwargs)

    Notes:
    ------
    - The function handles exceptions and returns error information in the `stats` dictionary.
    - The problem data (A, b, C) is fetched from the `get_problems()` function based on the index `i`.

    Raises:
    -------
    This function does not raise exceptions explicitly. Instead, it catches any exceptions and logs the error
    in the returned statistics dictionary.
    """
    fns = get_problems()
    A, b, C = fns[i]()
    initial_stats = initiate_stats(A)
    try:
        Z, X, RC, stats = multipivot_simplex(A, b, C, **kwargs)
        # Z, X, RC, stats = multipivot_simplex(A, b, C, **kwargs)
        return Z, np.all(X >= 0), np.max(RC), {**initial_stats, **stats}
    except Exception as e:
        error_message = traceback.format_exc()
        initial_stats["Message"] = f"{e} : {error_message}"
        return None, None, None, initial_stats


def save_to_csv(filename, problems, kwargs):
    """
    Executes the Multi-Pivot Simplex algorithm for multiple problem instances and saves the results to a CSV file.

    This function runs the `run_multipivot_simplex` function for problem instances indexed from 1 to 25,
    collects relevant statistics, and writes them to a specified CSV file. The header of the CSV file
    includes essential information about each problem's results.

    Parameters:
    ----------
    filename : str
        The name of the CSV file where the results will be saved. This file will be created or overwritten.
    kwargs : dict
        A dictionary of keyword arguments to pass to the `run_multipivot_simplex` function. These arguments
        customize the execution of the simplex algorithm for each problem instance (e.g., number of pivots,
        cut options, etc.).

    Returns:
    -------
    pandas.DataFrame
        A DataFrame containing the contents of the saved CSV file. This DataFrame includes all the results
        from the executed problems.

    Example:
    --------
    kwargs = {
        "Cut": True,
        "Order": False,
        "Pivots": 2,
    }
    df_results = save_to_csv("results.csv", kwargs)

    Notes:
    ------
    - The function runs the Multi-Pivot Simplex algorithm for 25 problem instances and writes
      the results, including the problem number, dimensions, objective values, feasibility,
      and various statistics to the specified CSV file.
    - The header of the CSV file includes columns for problem characteristics and performance metrics.
    - The function prints progress updates for each problem being saved.

    Raises:
    -------
    This function does not explicitly raise exceptions; however, it may encounter issues related to file writing
    or execution of the simplex algorithm, which should be handled by the caller.
    """
    header = [
        "Problem",
        "Cuts",
        "Inv",
        "Inf",
        "Nodes",
        "Nodes1",
        "Nodes2",
        "Nodes3",
        "Iter",
        "Time",
        "Message",
    ]
    with open(filename, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(header)
        for i in range(problems):
            print(f"Solving problem {i+1}")
            Z, Feasible, max_RC, stats = run_multipivot_simplex(i + 1, kwargs)
            message = stats["Message"]  # if Feasible else "Infeasible"
            row = [
                f"{i+1}",
                stats["NbrCuts"],
                stats["NbrIversions"],
                stats["NbrInfeasibilities"],
                stats["NbrNodes"],
                stats["NbrNodes_i"],
                stats["NbrNodes_j"],
                stats["NbrNodes_k"],
                stats["Iteration"],
                stats["Time"],
                message,
            ]
            writer.writerow(row)

    df = pd.read_csv(filename)
    return df


def save_dataframes_to_excel(file_name_excel, dfs):
    """
    Saves multiple pandas DataFrames to an Excel file with distinct sheet names.

    This function writes the provided DataFrames to separate sheets within the specified Excel file,
    allowing for organized data storage and easy access to different datasets.

    Parameters:
    ----------
    file_name_excel : str
        The name of the Excel file where the DataFrames will be saved, including the path and .xlsx extension.
    dfs : dict
        A dictionary where keys are the names of the sheets and values are the corresponding DataFrames to be saved.

    Returns:
    -------
    str
        The file path of the saved Excel file.

    Example:
    --------
    dataframes = {
        "MINE_CUT_ORDER": df_cut_order,
        "MINE_CUT_NO_ORDER": df_cut_no_order,
        "MINE_NO_CUT_ORDER": df_no_cut_order,
        "MINE_NO_CUT_NO_ORDER": df_no_cut_no_order,
    }

    file_path = save_dataframes_to_excel("results/result_file.xlsx", dataframes)
    print(f"Data saved to {file_path}")

    Notes:
    ------
    - The specified filename should include the desired path and ensure the directory exists.
    - If the directory does not exist, consider adding error handling to create it.

    Raises:
    -------
    FileNotFoundError: If the specified directory for the file does not exist.
    """
    # Save DataFrames to the specified Excel file
    df_list = []
    for sheet_name, df in dfs.items():
        df.insert(0, "Method", sheet_name)
        df_list.append(df)

    merged_df = pd.concat(df_list, ignore_index=True)
    count_optimal = merged_df[merged_df["Message"] == "Solved"].shape[0]
    cuts_max = merged_df["Cuts"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Cuts"] = cuts_max
    Inv_max = merged_df["Inv"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Inv"] = Inv_max
    Inf_max = merged_df["Inf"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Inf"] = Inf_max
    Nodes_max = merged_df["Nodes"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Nodes"] = Nodes_max
    Nodes1_max = merged_df["Nodes1"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Nodes1"] = Nodes1_max
    Nodes2_max = merged_df["Nodes2"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Nodes2"] = Nodes2_max
    Nodes3_max = merged_df["Nodes3"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Nodes3"] = Nodes3_max
    Iter_max = merged_df["Iter"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Iter"] = Iter_max
    Time_max = merged_df["Time"].max()
    merged_df.loc[merged_df["Message"] != "Solved", "Time"] = Time_max
    methdos = list(merged_df["Method"].unique())
    with pd.ExcelWriter(file_name_excel) as writer:
        merged_df.to_excel(writer, sheet_name="ALL", index=False)
        for method in methdos:
            df = merged_df.loc[merged_df["Method"] == method, :]
            df.to_excel(writer, sheet_name=method, index=False)

    wb = load_workbook(file_name_excel)
    ws = wb["ALL"]
    sum_row = len(merged_df) + 3
    ws[f"B{sum_row}"] = "Average"
    ws[f"C{sum_row}"] = f"=AVERAGE(C2:C{sum_row-1})"
    ws[f"D{sum_row}"] = f"=AVERAGE(D2:D{sum_row-1})"
    ws[f"E{sum_row}"] = f"=AVERAGE(E2:E{sum_row-1})"
    ws[f"F{sum_row}"] = f"=AVERAGE(F2:F{sum_row-1})"
    ws[f"G{sum_row}"] = f"=AVERAGE(G2:G{sum_row-1})"
    ws[f"H{sum_row}"] = f"=AVERAGE(H2:H{sum_row-1})"
    ws[f"I{sum_row}"] = f"=AVERAGE(I2:I{sum_row-1})"
    ws[f"J{sum_row}"] = f"=AVERAGE(J2:J{sum_row-1})"
    ws[f"K{sum_row}"] = f"=AVERAGE(K2:K{sum_row-1})"
    ws[f"L{sum_row}"] = f'=COUNTIF(L2:L{sum_row-1},"Solved")'
    ws[f"L{sum_row+1}"] = "Number Solved"
    for sheet_name, df in dfs.items():
        ws = wb[sheet_name]
        sum_row = len(df) + 3
        ws[f"B{sum_row}"] = "Average"
        ws[f"C{sum_row}"] = f"=AVERAGE(C2:C{sum_row-1})"
        ws[f"D{sum_row}"] = f"=AVERAGE(D2:D{sum_row-1})"
        ws[f"E{sum_row}"] = f"=AVERAGE(E2:E{sum_row-1})"
        ws[f"F{sum_row}"] = f"=AVERAGE(F2:F{sum_row-1})"
        ws[f"G{sum_row}"] = f"=AVERAGE(G2:G{sum_row-1})"
        ws[f"H{sum_row}"] = f"=AVERAGE(H2:H{sum_row-1})"
        ws[f"I{sum_row}"] = f"=AVERAGE(I2:I{sum_row-1})"
        ws[f"J{sum_row}"] = f"=AVERAGE(J2:J{sum_row-1})"
        ws[f"K{sum_row}"] = f"=AVERAGE(K2:K{sum_row-1})"
        ws[f"L{sum_row}"] = f'=COUNTIF(L2:L{sum_row-1},"Solved")'
        ws[f"L{sum_row+1}"] = "Number Solved"

    wb.save(file_name_excel)
    return count_optimal, len(merged_df), file_name_excel


if __name__ == "__main__":
    print("welcome to saveresults.py")


def save_results(problems=25, maxiters=1000):
    # print(py_example(1)[0])
    timestamp = datetime.datetime.now().strftime("%Y_%m_%d")
    # Define possible values for each keyword argument
    kwargs_options = {
        "Cut": [True, False],
        "Order": [True, False],
        "Pivots": [2, 3],
    }

    # Get all possible combinations of keyword argument values
    keys, values = zip(*kwargs_options.items())
    combinations = list(itertools.product(*values))
    dfs = {}
    for combo in combinations:
        kwargs = dict(
            zip(keys, combo)
        )  # Create a dictionary for the current combination

        # Create a descriptive name for the DataFrame
        df_name = f"df_C{kwargs['Cut']}_O{kwargs['Order']}_P{kwargs['Pivots']}"
        # Assuming save_to_csv function is defined and returns the DataFrame
        file_name_csv = f"./results/result_Cut_{kwargs['Cut']}_Order_{kwargs['Order']}_pivots_{kwargs['Pivots']}_{timestamp}.csv"
        print(f"Saving DataFrame for: {df_name} to {file_name_csv}")

        dfs[df_name] = save_to_csv(
            file_name_csv, problems, {"maxiters": maxiters, **kwargs}
        )

    file_name_excel = f"./results/result_{timestamp}.xlsx"

    count_solved, count_all, xfile_name_excel = save_dataframes_to_excel(
        file_name_excel, dfs
    )

    return count_solved, count_all, xfile_name_excel
